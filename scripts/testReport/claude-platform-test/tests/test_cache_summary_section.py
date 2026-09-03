"""汇总页的「④ 各模型缓存命中率汇总」：阈值跟随配置，两档都通过才算通过。"""

import pytest
from openpyxl import load_workbook

import config
from cases import caching
from cases.base import TestOutcome
from report import ResultRow, generate_report


def _cache_row(model, case_id, verdict, rate, read_tokens, creation_tokens):
    return ResultRow(
        model=model, case_id=case_id, category=caching.CATEGORY,
        name=caching.TTL_LABELS[case_id] + "缓存命中率", severity="P0",
        outcome=TestOutcome(
            verdict=verdict, expected="", actual="",
            metrics={
                caching.METRIC_HIT_RATE: rate,
                caching.METRIC_READ_TOKENS: read_tokens,
                caching.METRIC_CREATION_TOKENS: creation_tokens,
                caching.METRIC_FAILED_ROUNDS: 0,
            },
        ),
    )


def _summary_rows(tmp_path, rows, name="r.xlsx"):
    path = generate_report(rows, str(tmp_path / name))
    ws = load_workbook(path)["测试汇总"]
    return [[c for c in row] for row in ws.iter_rows(values_only=True)]


def _section(rows, marker):
    """取出某个段落的标题行和它下面的数据行。"""
    start = next(i for i, row in enumerate(rows)
                 if isinstance(row[0], str) and row[0].startswith(marker))
    out = [rows[start]]
    for row in rows[start + 1:]:
        if row[0] is None and all(v is None for v in row):
            break
        out.append(row)
    return out


@pytest.fixture
def _single_model(monkeypatch):
    monkeypatch.setattr(config, "MODELS", ["claude-opus-5"])
    return "claude-opus-5"


def test_threshold_follows_config(tmp_path, monkeypatch, _single_model):
    """调高 CACHE_HIT_PASS_RATIO，标题和「通过阈值」列都要跟着变。"""
    monkeypatch.setattr(config, "CACHE_HIT_PASS_RATIO", 0.9)
    rows = [
        _cache_row(_single_model, caching.CASE_5M, "PASS", "95.0%", 76000, 4000),
        _cache_row(_single_model, caching.CASE_1H, "PASS", "95.0%", 76000, 4000),
    ]

    section = _section(_summary_rows(tmp_path, rows), "④")

    assert "≥ 90% 判通过" in section[0][0]
    assert "读取 ÷ (读取 + 写入) tokens，含种子轮" in section[0][0]
    header, data = section[1], section[2]
    assert data[header.index("通过阈值")] == "90%"


def test_both_ttls_must_pass_for_the_overall_verdict(tmp_path, _single_model):
    rows = [
        _cache_row(_single_model, caching.CASE_5M, "FAIL", "70.0%", 56000, 24000),
        _cache_row(_single_model, caching.CASE_1H, "PASS", "95.0%", 76000, 4000),
    ]

    section = _section(_summary_rows(tmp_path, rows), "④")
    header, data = section[1], section[2]

    assert data[header.index("5分钟命中率")] == "70.0%"
    assert data[header.index("5分钟读取/写入tokens(含种子轮)")] == "56000 / 24000"
    assert data[header.index("5分钟结论")] == "不通过"
    assert data[header.index("1小时结论")] == "通过"
    assert data[header.index("综合结论")] == "不通过"


def test_all_pass_reports_overall_pass(tmp_path, _single_model):
    rows = [
        _cache_row(_single_model, caching.CASE_5M, "PASS", "100.0%", 80000, 0),
        _cache_row(_single_model, caching.CASE_1H, "PASS", "100.0%", 80000, 0),
    ]

    section = _section(_summary_rows(tmp_path, rows), "④")
    header, data = section[1], section[2]

    assert data[header.index("综合结论")] == "通过"


def test_missing_cache_rows_report_not_measured(tmp_path, _single_model):
    """模型没跑缓存用例时不能误报通过。"""
    rows = [ResultRow(_single_model, "AN-BASIC-001", "基础对话", "非流式基础对话", "P0",
                      TestOutcome(verdict="PASS"))]

    section = _section(_summary_rows(tmp_path, rows), "④")
    header, data = section[1], section[2]

    assert data[header.index("综合结论")] == "未测"


def test_sections_are_numbered_in_order(tmp_path, _single_model):
    rows = [
        _cache_row(_single_model, caching.CASE_5M, "PASS", "100.0%", 80000, 0),
        _cache_row(_single_model, caching.CASE_1H, "PASS", "100.0%", 80000, 0),
    ]

    titles = [row[0] for row in _summary_rows(tmp_path, rows)
              if isinstance(row[0], str) and row[0][:1] in "①②③④⑤⑥⑦⑧"]

    assert [t[0] for t in titles] == list("①①②③④⑤⑥⑦⑧")
    assert "缓存命中率汇总" in titles[4]
    assert "关键失败/异常清单" in titles[5]
    assert "Usage 聚合" in titles[6]
    assert "响应延迟分位" in titles[7]
    assert "结论摘要" in titles[8]
