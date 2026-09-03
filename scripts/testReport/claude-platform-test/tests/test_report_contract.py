from openpyxl import load_workbook

import config
from cases.base import TestOutcome
from report import ResultRow, generate_report


def _row():
    return ResultRow(
        model=config.MODELS[0],
        case_id="AN-BASIC-001",
        category="基础对话",
        name="非流式基础对话",
        severity="P0",
        outcome=TestOutcome(
            verdict="PASS",
            expected="返回文本",
            actual="ok",
            response_text="hello",
            response_id="msg_test_001",
            response_raw={"id": "msg_test_001", "type": "message"},
        ),
    )


def test_detail_and_summary_keep_only_the_required_report_sheets(tmp_path):
    path = generate_report([_row()], str(tmp_path / "report.xlsx"))
    workbook = load_workbook(path, read_only=True)

    assert workbook.sheetnames == ["测试汇总", "测试明细"]
    headers = list(next(workbook["测试明细"].iter_rows(values_only=True)))
    assert "模型预期Case" not in headers
    assert "Thinking模式" not in headers
    assert "模型系列" in headers


def test_detail_contains_response_id_and_raw_response(tmp_path):
    path = generate_report([_row()], str(tmp_path / "response.xlsx"))
    workbook = load_workbook(path, read_only=True)
    headers = list(next(workbook["测试明细"].iter_rows(values_only=True)))
    values = list(next(workbook["测试明细"].iter_rows(min_row=2, values_only=True)))
    by_name = dict(zip(headers, values))

    assert "msg_test_001" in by_name["实际结果（含模型完整响应）"]
    assert '"type": "message"' in by_name["实际结果（含模型完整响应）"]
