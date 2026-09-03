#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""报告输出路径与延迟表的"未采样"展示。

- generate_report 必须支持绝对路径（多任务并发时不能都写 CWD 互相覆盖）
- 没跑提示词缓存用例时，⑦ 表的分位数要写「未采样」而不是误导性的 0.0
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import config
from cases import latency as LATENCY
from cases.base import ok
from report import ResultRow, generate_report


def _one_row() -> ResultRow:
    return ResultRow(config.MODELS[0], "AN-BASIC-001", "基础对话", "非流式基础对话",
                     "P0", ok("有回复", "有回复", "通过"))


def test_generate_report_writes_to_absolute_path(tmp_path):
    target = tmp_path / "sub" / "报告.xlsx"
    written = generate_report([_one_row()], str(target))
    assert written == str(target)
    assert target.exists()


def test_latency_table_shows_not_sampled_when_no_samples(tmp_path):
    LATENCY.reset()
    target = tmp_path / "报告.xlsx"
    generate_report([_one_row()], str(target))

    ws = load_workbook(target).worksheets[0]
    texts = [str(row[0].value) for row in ws.iter_rows(min_col=1, max_col=1)]
    title_index = next(i for i, t in enumerate(texts) if t.startswith("⑦"))
    assert "无样本" in texts[title_index]

    # 表头行在标题下一行，数据行再下一行
    data_row = list(ws.iter_rows(min_row=title_index + 3, max_row=title_index + 3,
                                 min_col=1, max_col=9))[0]
    assert data_row[1].value == 0            # 样本数
    assert data_row[2].value == "未采样"      # P50
    assert data_row[7].value == "未采样"      # 最大


def test_latency_table_shows_numbers_when_sampled(tmp_path):
    LATENCY.reset()
    model = config.MODELS[0]
    for elapsed in (1.0, 2.0, 3.0):
        LATENCY.record(model, elapsed, succeeded=True)

    target = tmp_path / "报告2.xlsx"
    generate_report([_one_row()], str(target))

    ws = load_workbook(target).worksheets[0]
    texts = [str(row[0].value) for row in ws.iter_rows(min_col=1, max_col=1)]
    title_index = next(i for i, t in enumerate(texts) if t.startswith("⑦"))
    assert "无样本" not in texts[title_index]

    data_row = list(ws.iter_rows(min_row=title_index + 3, max_row=title_index + 3,
                                 min_col=1, max_col=9))[0]
    assert data_row[1].value == 3
    assert isinstance(data_row[2].value, (int, float))
    LATENCY.reset()
