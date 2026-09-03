#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试报告生成（Excel）
=====================

 Sheet 结构：
  1. 测试汇总 —— 整体通过率 / 分类 / 严重级别 / 缓存命中率 / 失败清单 / usage / 响应延迟分位
  2. 测试明细 —— 每行 = 一个模型 × 一个用例；包含上下文隔离详情和完整响应
"""

from __future__ import annotations

import datetime
import json
import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config
from cases import PASS, FAIL, SKIP, ERROR, TestOutcome
from cases import latency as LATENCY
from cases.caching import (
    TTL_LABELS as CACHE_TTL_LABELS,
    METRIC_HIT_RATE as CACHE_METRIC_HIT_RATE,
    METRIC_READ_TOKENS as CACHE_METRIC_READ,
    METRIC_CREATION_TOKENS as CACHE_METRIC_CREATION,
)
from cases.session_isolation import RECORDS as SESSION_ISOLATION_RECORDS

# 结论着色
_FILL = {
    PASS: PatternFill("solid", fgColor="C6EFCE"),   # 绿
    FAIL: PatternFill("solid", fgColor="FFC7CE"),   # 红
    SKIP: PatternFill("solid", fgColor="FFEB9C"),   # 黄
    ERROR: PatternFill("solid", fgColor="F4B084"),  # 橙
}
_FONT = {
    PASS: Font(color="006100"),
    FAIL: Font(color="9C0006"),
    SKIP: Font(color="9C6500"),
    ERROR: Font(color="843C0C"),
}
_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_SUBHEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
_VERDICT_CN = {PASS: "通过", FAIL: "不通过", SKIP: "跳过", ERROR: "异常"}

# ---- 排版参数 ----
_LINE_HEIGHT = 14.5          # 一行文字的高度（磅）
_ROW_PADDING = 3.0           # 行内上下留白
# 「测试汇总」各列宽度：A 放段落标题/模型名，G 放失败清单的关键原因，都要留足
_SUMMARY_WIDTHS = [57.2, 25.5, 16, 14, 10, 13, 48.7, 10, 16, 14]
# 「测试明细」各列宽度（模型/编号/分类/测试项/…/实际结果/判定说明/…/curl）
_DETAIL_WIDTHS = [26, 12, 15, 12, 26, 9, 12, 10, 30, 60, 34, 9, 22, 50]
_CN_VERDICT = {v: k for k, v in _VERDICT_CN.items()}


@dataclass
class ResultRow:
    model: str
    case_id: str
    category: str
    name: str
    severity: str
    outcome: TestOutcome


def _truncate_mid(text: str, limit: int | None = None) -> str:
    """超长文本省略中间。"""
    limit = config.EXCEL_CELL_MAX_CHARS if limit is None else limit
    if text is None:
        return ""
    text = str(text)
    if len(text) <= limit:
        return text
    head = limit * 2 // 3
    tail = limit - head
    return f"{text[:head]}\n...[省略中间 {len(text) - limit} 字符]...\n{text[-tail:]}"


def _truncate_tail(text: str, limit: int | None = None) -> str:
    """超长文本从尾部截断，加"...(已截断，原长 N)"提示。适合模型响应展示。"""
    limit = config.RESPONSE_TEXT_MAX_CHARS if limit is None else limit
    if text is None:
        return ""
    text = str(text)
    if len(text) <= limit:
        return text
    return f"{text[:limit]}\n...（已截断，原文共 {len(text)} 字符）"


# 这些字段是模型内部用的 base64 大 blob（思考签名、图片/PDF 数据、搜索结果密文），
# 原样写进 Excel 就是一屏乱码，对阅读报告的人没有任何价值，只保留长度说明。
_BLOB_KEYS = {"signature", "data", "encrypted_content", "encrypted_index"}
# JSON 字符串形态（流式原始 SSE 文本）里的同类字段
_BLOB_JSON_RE = re.compile(
    r'"(%s)"\s*:\s*"[^"]{40,}"' % "|".join(_BLOB_KEYS)
)


def _redact_blobs(value: Any) -> Any:
    """递归把 base64 大 blob 替换成简短说明，保持 JSON 结构可读。"""
    if isinstance(value, dict):
        out = {}
        for key, item in value.items():
            if key in _BLOB_KEYS and isinstance(item, str) and len(item) > 40:
                out[key] = f"<{key} base64 数据已省略，共 {len(item)} 字符>"
            else:
                out[key] = _redact_blobs(item)
        return out
    if isinstance(value, list):
        return [_redact_blobs(item) for item in value]
    return value


def _redact_blobs_text(text: str) -> str:
    """流式响应是原始 SSE 文本，按 JSON 字段名做同样的省略。"""
    return _BLOB_JSON_RE.sub(lambda m: f'"{m.group(1)}": "<base64 数据已省略>"', text)


def _metrics_and_usage(o: TestOutcome) -> str:
    """关键指标列：自定义指标 + 响应 usage 全量（usage 始终完整写入）。"""
    usage = o.response_usage or {}
    parts: list[str] = []
    # 自定义指标里剔除与 usage 重复的键，避免重复展示
    custom = {k: v for k, v in (o.metrics or {}).items() if k not in usage}
    if custom:
        parts.append("\n".join(f"{k}={v}" for k, v in custom.items()))
    if usage:
        parts.append("【usage】\n" + "\n".join(f"{k}={v}" for k, v in usage.items()))
    return "\n\n".join(parts)


def _format_response(o: TestOutcome) -> str:
    """把模型完整响应组织成"实际结果"列的多段展示。"""
    parts: list[str] = []
    # 摘要行（原 actual 字段）
    if o.actual:
        parts.append(f"【摘要】{o.actual}")
    # 主文本
    if o.response_text:
        parts.append(f"【模型响应文本】\n{_truncate_tail(o.response_text)}")
    if o.response_id:
        parts.append(f"【响应ID】{o.response_id}")
    if o.response_raw is not None:
        raw = (json.dumps(_redact_blobs(o.response_raw), ensure_ascii=False, indent=2)
               if isinstance(o.response_raw, (dict, list))
               else _redact_blobs_text(str(o.response_raw)))
        parts.append(f"【原始响应】\n{_truncate_mid(raw, limit=config.EXCEL_CELL_MAX_CHARS)}")
    # thinking
    if o.response_thinking:
        parts.append(f"【思考过程 thinking】\n{_truncate_tail(o.response_thinking)}")
    # 工具调用
    if o.response_tool_uses:
        tool_lines = []
        for tu in o.response_tool_uses:
            tool_lines.append(f"- {tu.get('name')}({tu.get('input')})")
        parts.append("【工具调用】\n" + "\n".join(tool_lines))
    if o.response_server_tool_uses:
        server_lines = []
        for tu in o.response_server_tool_uses:
            server_lines.append(f"- {tu.get('name')} ({tu.get('type')}) input={tu.get('input')}")
        parts.append("【服务端工具调用】\n" + "\n".join(server_lines))
    if o.response_server_tool_results:
        result_lines = []
        for result in o.response_server_tool_results:
            content = result.get("content") or []
            result_lines.append(f"- type={result.get('type')} 结果块数={len(content) if isinstance(content, list) else 1}")
        parts.append("【服务端工具结果】\n" + "\n".join(result_lines))
    # stop_reason
    if o.response_stop_reason:
        parts.append(f"【stop_reason】{o.response_stop_reason}")
    # 补充
    if o.response_extra:
        parts.append(f"【补充】{o.response_extra}")
    if not parts:
        return ""
    return "\n\n".join(parts)


def _column_units(text: Any) -> list[int]:
    """按 Excel 列宽单位估算每个换行段的宽度（西文 1 单位、中日韩 2 单位）。"""
    if text is None:
        return [0]
    return [sum(2 if ord(ch) > 0x2E80 else 1 for ch in segment)
            for segment in str(text).split("\n")]


def _wrapped_lines(text: Any, col_width: float) -> int:
    """估算一个单元格自动换行后占几行。"""
    usable = max(col_width - 1, 4)
    return sum(max(1, math.ceil(units / usable)) for units in _column_units(text))


def _fit_row_heights(ws, widths: list[float], max_height: float) -> None:
    """按内容估算并**显式写入**行高。

    openpyxl 不写行高时，Excel 打开会自行 autofit——超长单元格会被撑到 409.5
    的上限，整份报告每行一屏高，没法扫读。所以这里主动写死行高。
    """
    merged_spans = {(m.min_row, m.min_col): (m.min_col, m.max_col)
                    for m in ws.merged_cells.ranges}
    for row_cells in ws.iter_rows():
        lines = 1
        for cell in row_cells:
            if cell.value is None:
                continue
            span = merged_spans.get((cell.row, cell.column))
            if span:      # 合并单元格：可用宽度是整个跨列之和
                width = sum(widths[c - 1] for c in range(span[0], span[1] + 1)
                            if c - 1 < len(widths))
            else:
                width = widths[cell.column - 1] if cell.column - 1 < len(widths) else 10
            lines = max(lines, _wrapped_lines(cell.value, width))
        ws.row_dimensions[row_cells[0].row].height = min(
            max_height, _LINE_HEIGHT * lines + _ROW_PADDING)


def _style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_title(ws, row: int, text: str) -> int:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, size=12, color="1F4E78")
    cell.fill = _SUBHEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    return row + 1


# =============================================================================
# Sheet 1: 测试汇总（增强版）
# =============================================================================
def _write_summary_sheet(ws, rows: list[ResultRow]) -> None:
    ws.title = "测试汇总"
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ---- 元信息 ----
    r = 1
    ws.cell(row=r, column=1, value=f"Claude 资源测试报告 —— {config.PLATFORM_NAME}").font = Font(size=16, bold=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1
    for label, value in [
        ("生成时间", now),
        ("接口地址", f"{config.report_base_url()}/v1/messages"),
        ("鉴权方式", config.report_auth_mode()),
        ("测试模型", ", ".join(config.MODELS)),
        ("上下文隔离", "开启" if config.ENABLE_SESSION_ISOLATION else "关闭"),
        ("缓存命中率判定", f"种子轮 + {config.CACHE_HIT_ROUNDS} 轮读取；"
                          f"命中率 = 读取 ÷ (读取 + 写入) tokens（含种子轮，"
                          f"理论上限 {config.CACHE_HIT_ROUNDS / (config.CACHE_HIT_ROUNDS + 1):.1%}），"
                          f"≥ {config.CACHE_HIT_PASS_RATIO:.0%} 通过（详见 ④）"),
        ("并发线程数", config.MAX_WORKERS),
    ]:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=value)
        r += 1
    r += 1

    # ---- 1) 模型通过率对比 ----
    r = _write_title(ws, r, "① 各模型总体通过率")
    header = ["模型", "系列", "Thinking模式", "用例总数", "通过", "不通过", "异常", "跳过", "通过率(计入用例)", "整体健康度"]
    for c, h in enumerate(header, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    r += 1
    for model in config.MODELS:
        mrows = [x for x in rows if x.model == model]
        total = len(mrows)
        n_pass = sum(1 for x in mrows if x.outcome.verdict == PASS)
        n_fail = sum(1 for x in mrows if x.outcome.verdict == FAIL)
        n_err = sum(1 for x in mrows if x.outcome.verdict == ERROR)
        n_skip = sum(1 for x in mrows if x.outcome.verdict == SKIP)
        counted = total - n_skip
        rate = (n_pass / counted * 100) if counted else 0
        rate_str = f"{rate:.1f}%" if counted else "N/A"
        if not counted:
            health = "无有效用例"
        elif rate >= 90:
            health = "优秀"
        elif rate >= 70:
            health = "良好"
        elif rate >= 50:
            health = "及格"
        else:
            health = "需改进"
        vals = [model, config.model_series(model), config.thinking_mode(model), total,
                n_pass, n_fail, n_err, n_skip, rate_str, health]
        for c, val in enumerate(vals, start=1):
            ws.cell(row=r, column=c, value=val)
        r += 1
    r += 1

    # ---- 1.5) 按模型系列汇总 ----
    r = _write_title(ws, r, "①-2 按模型系列汇总")
    family_header = ["模型系列", "模型数", "用例总数", "通过", "不通过", "异常", "跳过", "通过率"]
    for c, h in enumerate(family_header, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    r += 1
    families: dict[str, list[ResultRow]] = {}
    for x in rows:
        families.setdefault(config.model_series(x.model), []).append(x)
    for family in sorted(families):
        items = families[family]
        total = len(items)
        n_pass = sum(x.outcome.verdict == PASS for x in items)
        n_fail = sum(x.outcome.verdict == FAIL for x in items)
        n_err = sum(x.outcome.verdict == ERROR for x in items)
        n_skip = sum(x.outcome.verdict == SKIP for x in items)
        counted = total - n_skip
        rate = f"{(n_pass / counted * 100):.1f}%" if counted else "N/A"
        vals = [family, len({x.model for x in items}), total, n_pass, n_fail, n_err, n_skip, rate]
        for c, val in enumerate(vals, start=1):
            ws.cell(row=r, column=c, value=val)
        r += 1
    r += 1

    # ---- 2) 按测试分类的通过率（跨所有模型合并） ----
    r = _write_title(ws, r, "② 按测试分类的通过率（跨所有模型合并）")
    cat_header = ["测试分类", "总数", "通过", "不通过", "异常", "跳过", "通过率"]
    for c, h in enumerate(cat_header, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    r += 1
    categories: dict[str, list[ResultRow]] = {}
    for x in rows:
        categories.setdefault(x.category, []).append(x)
    for cat, items in categories.items():
        total = len(items)
        n_pass = sum(1 for x in items if x.outcome.verdict == PASS)
        n_fail = sum(1 for x in items if x.outcome.verdict == FAIL)
        n_err = sum(1 for x in items if x.outcome.verdict == ERROR)
        n_skip = sum(1 for x in items if x.outcome.verdict == SKIP)
        counted = total - n_skip
        rate = f"{(n_pass / counted * 100):.1f}%" if counted else "N/A"
        for c, val in enumerate([cat, total, n_pass, n_fail, n_err, n_skip, rate], start=1):
            ws.cell(row=r, column=c, value=val)
        r += 1
    r += 1

    # ---- 3) 按严重级别的通过率 ----
    r = _write_title(ws, r, "③ 按严重级别的通过率（P0/P1/P2）")
    sev_header = ["严重级别", "总数", "通过", "不通过", "异常", "跳过", "通过率"]
    for c, h in enumerate(sev_header, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    r += 1
    severities: dict[str, list[ResultRow]] = {}
    for x in rows:
        severities.setdefault(x.severity or "-", []).append(x)
    for sev in sorted(severities.keys()):
        items = severities[sev]
        total = len(items)
        n_pass = sum(1 for x in items if x.outcome.verdict == PASS)
        n_fail = sum(1 for x in items if x.outcome.verdict == FAIL)
        n_err = sum(1 for x in items if x.outcome.verdict == ERROR)
        n_skip = sum(1 for x in items if x.outcome.verdict == SKIP)
        counted = total - n_skip
        rate = f"{(n_pass / counted * 100):.1f}%" if counted else "N/A"
        for c, val in enumerate([sev, total, n_pass, n_fail, n_err, n_skip, rate], start=1):
            ws.cell(row=r, column=c, value=val)
        r += 1
    r += 1

    # ---- 4) 各模型缓存命中率汇总 ----
    threshold = f"{config.CACHE_HIT_PASS_RATIO:.0%}"
    r = _write_title(
        ws, r,
        f"④ 各模型缓存命中率汇总（命中率 = 读取 ÷ (读取 + 写入) tokens，含种子轮，≥ {threshold} 判通过）")
    ttl_labels = list(CACHE_TTL_LABELS.values())
    cache_header = ["模型"]
    for label in ttl_labels:
        cache_header += [f"{label}命中率", f"{label}读取/写入tokens(含种子轮)", f"{label}结论"]
    cache_header += ["通过阈值", "综合结论"]
    for c, h in enumerate(cache_header, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    r += 1
    for model in config.MODELS:
        by_case = {x.case_id: x.outcome for x in rows if x.model == model}
        values: list[Any] = [model]
        verdicts: list[str] = []
        for case_id in CACHE_TTL_LABELS:
            outcome = by_case.get(case_id)
            if outcome is None:
                values += ["-", "-", "-"]
                continue
            m = outcome.metrics or {}
            verdict_cn = _VERDICT_CN.get(outcome.verdict, outcome.verdict)
            verdicts.append(outcome.verdict)
            values += [
                m.get(CACHE_METRIC_HIT_RATE, "-"),
                f"{m.get(CACHE_METRIC_READ, '-')} / {m.get(CACHE_METRIC_CREATION, '-')}",
                verdict_cn,
            ]
        # 两个 TTL 都通过才算整体通过；一个没测到就标"未测"
        if not verdicts:
            overall = "未测"
        elif all(v == PASS for v in verdicts):
            overall = _VERDICT_CN[PASS]
        else:
            overall = _VERDICT_CN[FAIL]
        values += [threshold, overall]
        verdict_columns = {1 + 3 * i + 3 for i in range(len(ttl_labels))} | {len(values)}
        for c, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if c in verdict_columns and val in _CN_VERDICT:
                cell.fill = _FILL[_CN_VERDICT[val]]
                cell.font = _FONT[_CN_VERDICT[val]]
                cell.alignment = Alignment(horizontal="center", vertical="center")
        r += 1
    r += 1

    # ---- 5) 关键失败清单（FAIL + ERROR） ----
    r = _write_title(ws, r, "⑤ 关键失败/异常清单（FAIL + ERROR）")
    fail_header = ["模型", "用例编号", "分类", "测试项", "严重", "结论", "关键原因"]
    for c, h in enumerate(fail_header, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    r += 1
    failed_rows = [x for x in rows if x.outcome.verdict in (FAIL, ERROR)]
    if not failed_rows:
        ws.cell(row=r, column=1, value="🎉 没有失败或异常用例").font = Font(color="006100", bold=True)
        r += 1
    else:
        for x in failed_rows:
            reason = (x.outcome.reason or x.outcome.actual or "")[:200]
            vals = [x.model, x.case_id, x.category, x.name, x.severity,
                    _VERDICT_CN.get(x.outcome.verdict, x.outcome.verdict), reason]
            for c, val in enumerate(vals, start=1):
                ws.cell(row=r, column=c, value=val)
            ws.cell(row=r, column=6).fill = _FILL.get(x.outcome.verdict, PatternFill())
            ws.cell(row=r, column=6).font = _FONT.get(x.outcome.verdict, Font())
            r += 1
    r += 1

    # ---- 6) 关键指标聚合（usage） ----
    r = _write_title(ws, r, "⑥ Usage 聚合（仅通过用例，跨所有模型合计）")
    metrics_header = ["模型", "累计 input_tokens", "累计 output_tokens", "累计 thinking_tokens",
                      "累计 cache_read", "累计 cache_creation", "有效样本数"]
    for c, h in enumerate(metrics_header, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    r += 1
    for model in config.MODELS:
        agg = {"input_tokens": 0, "output_tokens": 0, "thinking_tokens": 0,
               "cache_read_input_tokens": 0, "cache_creation_input_tokens": 0, "count": 0}
        for x in rows:
            if x.model != model or x.outcome.verdict != PASS:
                continue
            m = x.outcome.metrics or {}
            for k in ("input_tokens", "output_tokens",
                      "cache_read_input_tokens", "cache_creation_input_tokens"):
                v = m.get(k)
                if isinstance(v, int):
                    agg[k] += v
            details = m.get("output_tokens_details") or {}
            tt = details.get("thinking_tokens") if isinstance(details, dict) else None
            if isinstance(tt, int):
                agg["thinking_tokens"] += tt
            agg["count"] += 1
        vals = [model, agg["input_tokens"], agg["output_tokens"], agg["thinking_tokens"],
                agg["cache_read_input_tokens"], agg["cache_creation_input_tokens"], agg["count"]]
        for c, val in enumerate(vals, start=1):
            ws.cell(row=r, column=c, value=val)
        r += 1
    r += 1

    # ---- 7) 响应延迟分位 ----
    #   样本全部来自提示词缓存用例（见 cases/latency.py）。按分类筛选测试时
    #   如果没勾选「提示词缓存」，这里一个样本都没有，各分位写 0.0 会被误读成
    #   "延迟为零"，所以无样本时统一写「未采样」并在标题里点明原因。
    latency_sampled = any(LATENCY.stats(model)["样本数"] for model in config.MODELS)
    latency_title = ("⑦ 各模型响应延迟分位（取自提示词缓存用例的流式调用，"
                     "主指标为首字延迟，仅统计不判定）")
    if not latency_sampled:
        latency_title += "——本次未执行提示词缓存用例，无样本"
    r = _write_title(ws, r, latency_title)
    latency_header = ["模型", "样本数", "首字P50(s)", "首字P95(s)", "首字P99(s)",
                      "首字平均(s)", "总耗时P50(s)", "总耗时P95(s)", "失败调用数"]
    for c, h in enumerate(latency_header, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    r += 1
    for model in config.MODELS:
        data = LATENCY.stats(model)
        if data["样本数"]:
            vals = [model, data["样本数"], data["首字P50(s)"], data["首字P95(s)"],
                    data["首字P99(s)"], data["首字平均(s)"], data["总耗时P50(s)"],
                    data["总耗时P95(s)"], data["失败调用数"]]
        else:
            vals = [model, 0, "未采样", "未采样", "未采样",
                    "未采样", "未采样", "未采样", data["失败调用数"]]
        for c, val in enumerate(vals, start=1):
            ws.cell(row=r, column=c, value=val)
        r += 1
    r += 1

    # ---- 8) 结论文案 ----
    r = _write_title(ws, r, "⑧ 结论摘要")
    n_pass_all = sum(1 for x in rows if x.outcome.verdict == PASS)
    n_fail_all = sum(1 for x in rows if x.outcome.verdict == FAIL)
    n_err_all = sum(1 for x in rows if x.outcome.verdict == ERROR)
    n_skip_all = sum(1 for x in rows if x.outcome.verdict == SKIP)
    total_all = len(rows)
    valid = total_all - n_skip_all
    overall_rate = (n_pass_all / valid * 100) if valid else 0
    conclusion_lines = [
        f"共执行 {total_all} 项 (模型 × 用例)，其中有效 {valid} 项，跳过 {n_skip_all} 项。",
        f"通过 {n_pass_all} / 不通过 {n_fail_all} / 异常 {n_err_all}，整体通过率 {overall_rate:.1f}%。",
    ]
    if failed_rows:
        conclusion_lines.append(
            f"关键失败项集中在：{', '.join(sorted({x.category for x in failed_rows}))}。请见 ⑤ 明细。"
        )
    latency_line = "、".join(
        f"{model} 首字P50={LATENCY.stats(model)['首字P50(s)']}s"
        f"/P95={LATENCY.stats(model)['首字P95(s)']}s"
        for model in config.MODELS if LATENCY.stats(model)["样本数"]
    )
    if latency_line:
        conclusion_lines.append(f"⏱ 流式首字延迟（仅统计不计通过率）：{latency_line}。详见 ⑦。")
    if SESSION_ISOLATION_RECORDS:
        leaked = [r for r in SESSION_ISOLATION_RECORDS if not r.overall_pass]
        if leaked:
            conclusion_lines.append(
                f"⚠️ 上下文隔离存在风险：{len(leaked)}/{len(SESSION_ISOLATION_RECORDS)} 个模型存在自身召回失败或跨会话泄露，详见测试明细中的 AN-SEC-001。"
            )
        else:
            conclusion_lines.append(
                f"✅ 上下文隔离测试通过：{len(SESSION_ISOLATION_RECORDS)} 个模型全部符合预期，详见测试明细中的 AN-SEC-001。"
            )
    for line in conclusion_lines:
        ws.cell(row=r, column=1, value=line)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1

    for i, w in enumerate(_SUMMARY_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row_cells in ws.iter_rows():
        for cell in row_cells:
            if cell.alignment.horizontal is None:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    _fit_row_heights(ws, _SUMMARY_WIDTHS, config.EXCEL_SUMMARY_ROW_MAX_HEIGHT)


# =============================================================================
# Sheet 2: 测试明细（带模型完整响应）
# =============================================================================
def _write_detail_sheet(ws, rows: list[ResultRow]) -> None:
    ws.title = "测试明细"
    headers = ["模型", "模型系列", "用例编号", "测试分类", "测试项", "严重级别",
               "边界预期", "测试结论", "预期结果", "实际结果（含模型完整响应）", "判定说明",
               "HTTP状态", "关键指标", "调用样例(curl)"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for row in rows:
        o = row.outcome
        actual_display = _format_response(o)
        boundary_expectation = "预期4xx" if o.metrics.get("expected_status") == "4xx" else ""
        ws.append([
            row.model,
            config.model_series(row.model),
            row.case_id,
            row.category,
            row.name,
            row.severity,
            boundary_expectation,
            _VERDICT_CN.get(o.verdict, o.verdict),
            _truncate_mid(o.expected),
            _truncate_mid(actual_display, limit=config.EXCEL_CELL_MAX_CHARS),
            _truncate_mid(o.reason),
            o.status_code if o.status_code is not None else "",
            _metrics_and_usage(o),
            # curl 尽量完整（含超长系统提示词/缓存标记），仅在逼近 Excel 单元格上限时才省略中间
            _truncate_mid(o.curl, limit=30000),
        ])
        # 给"测试结论"列(第8列)着色
        cell = ws.cell(row=ws.max_row, column=8)
        cell.fill = _FILL.get(o.verdict, PatternFill())
        cell.font = _FONT.get(o.verdict, Font())
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, w in enumerate(_DETAIL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    # 表头按内容估高；数据行固定行高——"实际结果"列动辄上千字符，
    # 不写死行高的话 Excel 会 autofit 到 409.5 的上限，一行占满一屏没法扫读。
    # 想看全文可以单独拖高那一行，或点单元格在编辑栏里读。
    _fit_row_heights(ws, _DETAIL_WIDTHS, 40)
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = config.EXCEL_DETAIL_ROW_HEIGHT
    ws.freeze_panes = "A2"


# =============================================================================
# 入口
# =============================================================================
def generate_report(rows: list[ResultRow], filename: str | None = None) -> str:
    """生成 Excel 报告，返回实际写入的路径。

    filename 传绝对路径时写到该路径（父目录不存在会自动创建）；
    留空则沿用 config.REPORT_FILENAME 写当前工作目录，保持手工执行的老行为。
    多任务并发跑测试时必须传绝对路径，否则文件名相同会互相覆盖。
    """
    filename = filename or config.REPORT_FILENAME
    target = Path(filename)
    if not target.parent.exists():
        target.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    _write_summary_sheet(wb.active, rows)
    _write_detail_sheet(wb.create_sheet(), rows)
    wb.save(str(target))
    return str(target)
